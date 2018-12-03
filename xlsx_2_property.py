#!/usr/bin/python3.4

import sys
import openpyxl
import subprocess
import logging
import pdb

logging.basicConfig(filename='app_xlsx_2_property.log', level=logging.DEBUG, filemode='w',
                    format='%(asctime)s - %(levelname)s - %(message)s')

logging.info('Otvaram datoteku "' + sys.argv[1] + '" .....')
wb = openpyxl.load_workbook(sys.argv[1])

logging.info('Lista sheet-ova: ' + str(wb.sheetnames))

for s in wb.sheetnames[1:]:

    sheet = wb[s]

    # Definiraje putanje do property datoteka
    if sheet.title[:4] == 'CCMS':
        path = 'D3/ccms'
    elif sheet.title[:4] == 'MMPA':
        path = 'D3/mmpa'
    else:
        logging.critical(sheet.title[:4] + ' je neispravno ime sheet-a !!! Prekidam izvođenje skripte.')
        break

    # Definiranje reda u kojem je vrijednost. rowValue vrijednost je broj koji se zbraja sa brojem reda u kojem je ime placeholder
    if sys.argv[2] == 't1':
        rowValue = 0
    elif sys.argv[2] == 'm1':
        rowValue = 1
    elif sys.argv[2] == 'r1':
        rowValue = 2
    elif sys.argv[2] == 'd3':
        rowValue = 3
    else:
        print('Ne postoji okolina "' + sys.argv[2] + '" !!!\n')
        logging.critical('Ne postoji okolina "' + sys.argv[2] + '" !!! Prekidam sa radom!')
        break

    logging.info('\n------------------\n' + sheet.title + '\n------------------\n')

    for row in range(2, sheet.max_row + 1):
        if sheet['D' + str(row)].value is not None:

            ph_name = sheet['D' + str(row)].value
            ph_value = ''

            if sheet['I' + str(row + rowValue)].value is not None:
                ph_value = str(sheet['I' + str(row + rowValue)].value)
            elif sheet['J' + str(row + rowValue)].value is not None:
                ph_value = str(sheet['J' + str(row + rowValue)].value)
            elif sheet['K' + str(row + rowValue)].value is not None:
                ph_value = str(sheet['K' + str(row + rowValue)].value)
            else:
                logging.error(
                    'Za placeholder "' + ph_name + '" iz sheet-a "' + sheet.title + '" ne postoji vrijednost u tablici !!!"')

            try:
                p1 = subprocess.Popen(("grep", "-R", "%s" % ph_name, path), stdout=subprocess.PIPE)
                file_list = p1.communicate()[0].splitlines()
            except:
                logging.error('Ne mogu radit "grep" na putanji "' + path + '" !!!')

            if len(file_list) > 0:

                for config_file in file_list:
                    try:
                        # ovdje treba return od config_file castati u string,jer je tip bytes, pa split() u nastavku nebi radio ako se ne casta iz bytes u string (to je novost u python3,a u python2 to radi bez castanja)
                        config_file_path = str(config_file).split(':')[0].split("'")[1]
                        with open(config_file_path, 'r') as f:
                            newlines = []
                            for line in f.readlines():
                                newlines.append(line.replace(ph_name, ph_value))
                        with open(config_file_path, 'w') as f:
                            for line in newlines:
                                f.write(line)
                    except FileNotFoundError:
                        logging.error(
                            'Ne mogu otvoriti datoteku u kojoj je placeholder "' + ph_name + '"')
                        continue
                    # U Pythonu se moze koristit else u try: bloku, da se nesto ispise, ako je try blok uspjesno odradjen
                    else:
                        logging.info(
                            'Placeholder "' + ph_name + '" iz sheet-a "' + sheet.title + '" popunjen u datoteci "' + config_file_path + '" sa vrijednoscu "' + ph_value + '" !')

            else:
                logging.error(
                    'Placeholder "' + ph_name + '" iz sheet-a "' + sheet.title + '" ne postoji u datotekama na putanji "' + path + '"')

logging.info('\n------------------\nEXTRA_properties\n------------------\n')

try:
    with open('ccmsCCUK.properties.build.EXTRA.txt', 'r') as f:
        newlines = []
        for line in f.readlines():
            newlines.append(line)
    with open('D3/ccms/config-CU/properties.build', 'a') as f:
        for line in newlines:
            f.write(line)
    logging.info(
        'Sadrzaj datoteke "ccmsCCUK.properties.build.EXTRA.txt" dodan u datoteku "D3/ccms/config-CU/properties.build"')

    with open('ccmsCCEU.properties.build.EXTRA.txt', 'r') as f:
        newlines = []
        for line in f.readlines():
            newlines.append(line)
    with open('D3/ccms/config-CE/properties.build', 'a') as f:
        for line in newlines:
            f.write(line)
    logging.info(
        'Sadrzaj datoteke "ccmsCCEU.properties.build.EXTRA.txt" dodan u datoteku "D3/ccms/config-CE/properties.build"')
except FileNotFoundError:
    logging.error('Ne mogu otvoriti datoteku prilikom dodavanja extra property-a !!!')

sys.exit(1)